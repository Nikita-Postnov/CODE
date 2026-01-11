import imaplib
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import json
import os
import sys
import base64
import re
from datetime import datetime, timedelta
from email.utils import parsedate_to_datetime
import threading
import traceback
import logging
import hashlib
import mimetypes
import smtplib
from email.message import EmailMessage
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

BASE_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
LOG_PATH = os.path.join(BASE_DIR, "kolovrat.log")
REPORTS_DIR = os.path.join(BASE_DIR, "Отчеты")

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

def _safe_filename(s: str, max_len: int = 80) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^0-9A-Za-zА-Яа-я._-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        s = "report"
    return s[:max_len]

def save_full_report(
    title: str,
    message: str,
    username: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
) -> str | None:
    try:
        os.makedirs(REPORTS_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        user_part = _safe_filename(username or "account", 40)
        period_part = _safe_filename(f"{from_date or ''}_{to_date or ''}", 30)
        name = f"Отчет_{user_part}_{period_part}_{ts}.txt"
        out_path = os.path.join(REPORTS_DIR, name)
        with open(out_path, "w", encoding="utf-8") as f:
            f.write((title or "Результат") + "\n\n")
            f.write(message or "")
            f.write("\n")
        return out_path
    except Exception:
        logging.exception("Не удалось сохранить отчёт")
        return None
    
def save_excel_report(
    username: str,
    from_date: str,
    to_date: str,
    total: int,
    inbox: int,
    sent: int,
) -> str | None:
    try:
        os.makedirs(REPORTS_DIR, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        user_part = _safe_filename(username or "account", 40)
        period_part = _safe_filename(f"{from_date}_{to_date}", 30)
        filename = f"Отчет_{user_part}_{period_part}_{ts}.xlsx"
        path = os.path.join(REPORTS_DIR, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        ws.append(["Параметр", "Значение"])
        ws.append(["Аккаунт", username])
        ws.append(["Период", f"с {from_date} по {to_date}"])
        ws.append(["Всего писем", total])
        ws.append(["Входящие", inbox])
        ws.append(["Исходящие", sent])
        ws.append(["Дата формирования", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        for col in ("A", "B"):
            ws.column_dimensions[col].width = 30

        wb.save(path)
        return path
    except Exception:
        logging.exception("Не удалось сохранить Excel-отчёт")
        return None
    
def save_common_excel_report(
    username: str,
    from_date: str,
    to_date: str,
    total: int,
    inbox: int,
    sent: int,
) -> str | None:
    try:
        os.makedirs(REPORTS_DIR, exist_ok=True)
        path = os.path.join(REPORTS_DIR, "Общий отчет.xlsx")
        thin = Side(style="thin", color="000000")
        sep = Side(style="medium", color="000000")
        border_A_mid = Border(right=thin)
        border_B_mid = Border(left=thin)
        header_A = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_B = Border(left=thin, right=thin, top=thin, bottom=thin)
        bottom_A = Border(right=thin, bottom=sep)
        bottom_B = Border(left=thin, bottom=sep)
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Общий отчет"
            ws.append(["Параметр", "Значение"])
            ws["A1"].border = header_A
            ws["B1"].border = header_B
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 40
        if ws["A1"].value == "Параметр" and ws["B1"].value == "Значение":
            ws["A1"].border = header_A
            ws["B1"].border = header_B
        start_row = ws.max_row + 1
        ws.append(["Аккаунт", username])
        ws.append(["Период", f"с {from_date} по {to_date}"])
        ws.append(["Всего писем", total])
        ws.append(["Входящие", inbox])
        ws.append(["Исходящие", sent])
        ws.append(["Дата формирования", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        end_row = ws.max_row

        def merge_border(cell, *, left=None, right=None, top=None, bottom=None):
            b = cell.border
            cell.border = Border(
                left=left or b.left,
                right=right or b.right,
                top=top or b.top,
                bottom=bottom or b.bottom,
            )

        for r in range(start_row, end_row + 1):
            merge_border(ws[f"A{r}"], right=thin)
            merge_border(ws[f"B{r}"], left=thin, right=thin)
        merge_border(ws[f"A{end_row}"], bottom=sep)
        merge_border(ws[f"B{end_row}"], bottom=sep, right=thin)
        wb.save(path)
        return path
    except Exception:
        logging.exception("Не удалось сохранить общий Excel-отчёт")
        return None

def ensure_required_files():
    p = os.path.join(BASE_DIR, "mailsettings.json")
    if not os.path.exists(p):
        with open(p, "w", encoding="utf-8") as f:
            f.write("{}")
    accounts_path = os.path.join(BASE_DIR, "accounts.csv")
    if not os.path.exists(accounts_path):
        with open(accounts_path, "w", encoding="utf-8") as f:
            f.write("# логин:пароль\n")
    hosts_path = os.path.join(BASE_DIR, "imap_hosts.csv")
    if not os.path.exists(hosts_path):
        with open(hosts_path, "w", encoding="utf-8") as f:
            f.write("# host:port\n")
            f.write("imap.yandex.ru:993\n")
            f.write("mail.klvrt.ru:993\n")

def to_imap_date(date_str: str) -> str:
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{dt.day:02d}-{months[dt.month - 1]}-{dt.year}"

def connect_imap(host: str, port: int, username: str, password: str, timeout: float = 30.0) -> imaplib.IMAP4_SSL:
    mail = imaplib.IMAP4_SSL(host, port, timeout=timeout)
    mail.login(username, password)
    setattr(mail, "_kolovrat_host", host)
    return mail

def fetch_message_ids(mail, uid_list, step: int = 800) -> set[str]:
    if not uid_list:
        return set()
    res: set[str] = set()
    norm_uids: list[bytes] = []
    for u in uid_list:
        if u is None:
            continue
        if isinstance(u, bytes):
            norm_uids.append(u)
        else:
            norm_uids.append(str(u).encode("ascii", "ignore"))
    uid_re = re.compile(rb"\bUID\s+(\d+)\b", re.IGNORECASE)
    for i in range(0, len(norm_uids), step):
        chunk = norm_uids[i:i + step]
        uid_set = b",".join(chunk)
        status, data = mail.uid(
            "FETCH",
            uid_set,
            "(UID BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)])"
        )
        if status != "OK" or not data:
            continue
        for item in data:
            if not item or not isinstance(item, tuple) or len(item) < 2:
                continue
            meta, payload = item[0], item[1]
            if not isinstance(payload, (bytes, bytearray)):
                continue
            uid = None
            try:
                meta_b = meta if isinstance(meta, (bytes, bytearray)) else str(meta).encode("utf-8", "ignore")
                m = uid_re.search(meta_b)
                if m:
                    uid = m.group(1).decode("ascii", "ignore")
            except Exception:
                uid = None
            mid = ""
            try:
                lines = bytes(payload).splitlines()
                for idx, line in enumerate(lines):
                    if line.lower().startswith(b"message-id:"):
                        val = line.split(b":", 1)[1].strip()
                        j = idx + 1
                        while j < len(lines) and (lines[j].startswith(b" ") or lines[j].startswith(b"\t")):
                            val += b" " + lines[j].strip()
                            j += 1
                        val = val.strip().strip(b"<").strip(b">").strip()
                        mid = val.decode("utf-8", "ignore").strip()
                        break
            except Exception:
                mid = ""
            if mid:
                res.add(mid)
            elif uid:
                res.add("uid:" + uid)
            else:

                pass
    return res

def fetch_message_id_map(mail, uid_list, step: int = 800) -> dict[str, str]:
    if not uid_list:
        return {}
    out: dict[str, str] = {}
    norm_uids: list[bytes] = []
    for u in uid_list:
        if u is None:
            continue
        if isinstance(u, bytes):
            norm_uids.append(u)
        else:
            norm_uids.append(str(u).encode("ascii", "ignore"))
    uid_re = re.compile(rb"\bUID\s+(\d+)\b", re.IGNORECASE)
    for i in range(0, len(norm_uids), step):
        chunk = norm_uids[i:i + step]
        uid_set = b",".join(chunk)
        status, data = mail.uid(
            "FETCH",
            uid_set,
            "(UID BODY.PEEK[HEADER.FIELDS (MESSAGE-ID)])"
        )
        if status != "OK" or not data:
            continue
        for item in data:
            if not item or not isinstance(item, tuple) or len(item) < 2:
                continue
            meta, payload = item[0], item[1]
            if not isinstance(payload, (bytes, bytearray)):
                continue
            uid = ""
            try:
                meta_b = meta if isinstance(meta, (bytes, bytearray)) else str(meta).encode("utf-8", "ignore")
                m = uid_re.search(meta_b)
                if m:
                    uid = m.group(1).decode("ascii", "ignore").strip()
            except Exception:
                uid = ""
            mid = ""
            try:
                lines = bytes(payload).splitlines()
                for idx, line in enumerate(lines):
                    if line.lower().startswith(b"message-id:"):
                        val = line.split(b":", 1)[1].strip()
                        j = idx + 1
                        while j < len(lines) and (lines[j].startswith(b" ") or lines[j].startswith(b"\t")):
                            val += b" " + lines[j].strip()
                            j += 1
                        val = val.strip().strip(b"<").strip(b">").strip()
                        mid = val.decode("utf-8", "ignore").strip()
                        break
            except Exception:
                mid = ""
            if uid:
                out[uid] = mid if mid else ("uid:" + uid)
    return out

def _load_mid_cache(path: str) -> dict:
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                obj = json.load(f)
            if isinstance(obj, dict):
                return obj
    except Exception:
        pass
    return {}

def _save_mid_cache(path: str, obj: dict) -> None:
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False)
    except Exception:
        pass

def _get_uidvalidity(mail, encoded_folder: str) -> str:
    safe_name = encoded_folder.replace("\\", "\\\\").replace('"', r'\"')
    try:
        status, data = mail.status(f'"{safe_name}"', "(UIDVALIDITY)")
        if status == "OK" and data:
            blob = b" ".join(x for x in data if isinstance(x, (bytes, bytearray)))
            m = re.search(rb"UIDVALIDITY\s+(\d+)", blob)
            if m:
                return m.group(1).decode("ascii", "ignore")
    except Exception:
        pass
    return ""

def _replace_sent_with_internal(criteria: list[str]) -> list[str]:
    alt = list(criteria)
    for i, v in enumerate(alt):
        try:
            up = v.upper()
        except Exception:
            continue
        if up == "SENTSINCE":
            alt[i] = "SINCE"
        elif up == "SENTBEFORE":
            alt[i] = "BEFORE"
    return alt

def search_uids(mail, encoded_folder, from_date, to_date, extra_criteria=None, include_deleted=False, use_sent_date=False):
    safe_name = encoded_folder.replace("\\", "\\\\").replace('"', r"\"")
    status, _ = mail.select(f'"{safe_name}"', readonly=True)
    if status != "OK":
        raise RuntimeError("Не удалось открыть одну из папок почты. Проверь доступ к папкам и попробуй ещё раз.")
    criteria: list[str] = ["ALL" if include_deleted else "UNDELETED"]
    dt_to = datetime.strptime(to_date, "%Y-%m-%d") + timedelta(days=1)
    if use_sent_date:
        criteria.extend(["SENTSINCE", to_imap_date(from_date), "SENTBEFORE", to_imap_date(dt_to.strftime("%Y-%m-%d"))])
    else:
        criteria.extend(["SINCE", to_imap_date(from_date), "BEFORE", to_imap_date(dt_to.strftime("%Y-%m-%d"))])
    if extra_criteria:
        criteria.extend(extra_criteria)
    status, data = mail.uid("SEARCH", *criteria)
    if status != "OK":
        if use_sent_date:
            try:
                alt = _replace_sent_with_internal(criteria)
                status2, data2 = mail.uid("SEARCH", *alt)
                if status2 == "OK" and data2 and data2[0]:
                    return data2[0].split()
            except Exception:
                pass
        raise RuntimeError("Не удалось выполнить поиск писем в одной из папок. Попробуй повторить расчёт или уменьшить период.")
    result = data[0].split() if data and data[0] else []
    if use_sent_date and not result:
        try:
            alt = _replace_sent_with_internal(criteria)
            status2, data2 = mail.uid("SEARCH", *alt)
            if status2 == "OK" and data2 and data2[0]:
                result = data2[0].split()
        except Exception:
            pass
    return result

def decode_imap_utf7(s: str) -> str:
    res: list[str] = []
    i = 0
    while i < len(s):
        c = s[i]
        if c == "&":
            j = s.find("-", i + 1)
            if j == -1:
                res.append(s[i:])
                break
            if j == i + 1:
                res.append("&")
                i = j + 1
                continue
            b64 = s[i + 1: j].replace(",", "/")
            try:
                pad = "=" * (-len(b64) % 4)
                data = base64.b64decode(b64 + pad)
                res.append(data.decode("utf-16-be"))
            except Exception:
                res.append(s[i: j + 1])
            i = j + 1
        else:
            res.append(c)
            i += 1
    return "".join(res)


class MailStatsUI(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("КОЛОВРАТ")
        self.resizable(False, False)
        self._date_guard = False
        self._config_path = os.path.join(BASE_DIR, "mailsettings.json")
        self._accounts_path = os.path.join(BASE_DIR, "accounts.csv")
        self._imap_hosts_path = os.path.join(BASE_DIR, "imap_hosts.csv")
        self._accounts: dict[str, str] = {}
        self._imap_hosts: dict[str, int] = {}
        self._load_accounts()
        self._load_imap_hosts()
        self._build_ui()
        self._load_config()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self._worker_thread: threading.Thread | None = None
        self._cancel_event = threading.Event()
        self._busy = False
        self._last_report_files: list[str] = []

    def _load_accounts(self) -> None:
        self._accounts.clear()
        if not os.path.exists(self._accounts_path):
            return
        try:
            with open(self._accounts_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if ":" not in line:
                        continue
                    login, password = line.split(":", 1)
                    login = login.strip()
                    password = password.strip()
                    if login and password:
                        self._accounts[login] = password
        except Exception as e:
            print(f"Ошибка загрузки аккаунтов: {e}")

    def _get_selected_accounts(self) -> list[str]:
        try:
            if not hasattr(self, "accounts_listbox"):
                return []
            idx = list(self.accounts_listbox.curselection())
            if not idx:
                return []
            out: list[str] = []
            for i in idx:
                v = (self.accounts_listbox.get(i) or "").strip()
                if v:
                    out.append(v)
            seen: set[str] = set()
            uniq: list[str] = []
            for x in out:
                if x in seen:
                    continue
                seen.add(x)
                uniq.append(x)
            return uniq
        except Exception:
            return []

    def _refresh_accounts_listbox(self) -> None:
        if not hasattr(self, "accounts_listbox"):
            return
        try:
            selected = set(self._get_selected_accounts())
            self.accounts_listbox.delete(0, tk.END)
            vals = sorted(self._accounts.keys())
            for acc in vals:
                self.accounts_listbox.insert(tk.END, acc)
            for i, acc in enumerate(vals):
                if acc in selected:
                    self.accounts_listbox.selection_set(i)
        except Exception:
            pass

    def _load_mailsettings(self) -> dict:
        try:
            if os.path.exists(self._config_path):
                with open(self._config_path, "r", encoding="utf-8") as f:
                    obj = json.load(f)
                return obj if isinstance(obj, dict) else {}
        except Exception:
            pass
        return {}

    def _save_mailsettings(self, obj: dict) -> None:
        try:
            with open(self._config_path, "w", encoding="utf-8") as f:
                json.dump(obj, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _is_archive_folder(self, flags_part: str, decoded_name: str) -> bool:
        flags_upper = (flags_part or "").upper()
        name_upper = (decoded_name or "").upper()
        if "\\ARCHIVE" in flags_upper:
            return True
        archive_keywords = [
            "ARCHIVE",
            "ARCHIVES",
            "АРХИВ",
            "ARCHIV",
        ]
        for kw in archive_keywords:
            if kw in name_upper:
                return True
        return False
    
    def on_send_reports(self) -> None:
        files = [p for p in (self._last_report_files or []) if p and os.path.exists(p)]
        if not files:
            messagebox.showwarning("Внимание", "Нет файлов отчётов для отправки. Сначала выполни анализ.", parent=self)
            return
        self._open_send_dialog(files)

    def _open_send_dialog(self, files: list[str]) -> None:
        dlg = tk.Toplevel(self)
        dlg.transient(self)
        dlg.title("Отправка отчётов")
        dlg.resizable(False, False)
        dlg.grab_set()
        frm = ttk.Frame(dlg, padding=10)
        frm.grid(row=0, column=0, sticky="nsew")
        ttk.Label(frm, text="С какого e-mail отправлять (From):").grid(row=0, column=0, sticky="w")
        from_var = tk.StringVar()
        entry_from = ttk.Entry(frm, textvariable=from_var, width=45)
        entry_from.grid(row=1, column=0, sticky="we", pady=(0, 8))
        ttk.Label(frm, text="На какой e-mail отправлять (To):").grid(row=2, column=0, sticky="w")
        to_var = tk.StringVar()
        entry_to = ttk.Entry(frm, textvariable=to_var, width=45)
        entry_to.grid(row=3, column=0, sticky="we", pady=(0, 8))
        ttk.Label(frm, text="SMTP хост (сервер исходящей почты):").grid(row=4, column=0, sticky="w")
        smtp_host_var = tk.StringVar()
        combo_smtp_host = ttk.Combobox(frm, textvariable=smtp_host_var, width=45, state="normal")
        combo_smtp_host.grid(row=5, column=0, sticky="we", pady=(0, 8))
        ttk.Label(frm, text="SMTP порт:").grid(row=6, column=0, sticky="w")
        smtp_port_var = tk.StringVar()
        entry_smtp_port = ttk.Entry(frm, textvariable=smtp_port_var, width=10)
        entry_smtp_port.grid(row=7, column=0, sticky="w", pady=(0, 8))
        ttk.Label(frm, text="Выбери файлы для отправки (можно несколько):").grid(row=8, column=0, sticky="w")
        lb = tk.Listbox(frm, height=6, selectmode=tk.EXTENDED, exportselection=False, width=80)
        lb.grid(row=9, column=0, sticky="we")
        for p in files:
            lb.insert(tk.END, p)
        lb.selection_set(0, tk.END)
        preview = tk.Text(frm, height=6, width=80)
        preview.grid(row=10, column=0, sticky="we", pady=(8, 8))
        preview.config(state="disabled")

        def update_preview(event=None):
            try:
                sel = list(lb.curselection())
                chosen = [lb.get(i) for i in sel] if sel else []
            except Exception:
                chosen = []
            preview.config(state="normal")
            preview.delete("1.0", tk.END)
            if chosen:
                preview.insert(tk.END, "Будет отправлено:\n" + "\n".join(chosen))
            else:
                preview.insert(tk.END, "Ничего не выбрано.")
            preview.config(state="disabled")

        lb.bind("<<ListboxSelect>>", update_preview)
        update_preview()
        settings = self._load_mailsettings()
        if not isinstance(settings, dict):
            settings = {}
        saved_smtp_host = (settings.get("smtp_host") or "").strip()
        saved_smtp_port = settings.get("smtp_port")
        last_from = (settings.get("smtp_from") or settings.get("last_report_from") or "").strip()
        last_to = (settings.get("smtp_to") or settings.get("last_report_to") or "").strip()
        if last_from:
            from_var.set(last_from)
        if last_to:
            to_var.set(last_to)
        last_pwd_plain: str | None = None
        encoded_pwd = settings.get("last_report_password_b64")
        if isinstance(encoded_pwd, str) and encoded_pwd.strip():
            try:
                last_pwd_plain = base64.b64decode(encoded_pwd.encode("ascii")).decode("utf-8")
            except Exception:
                last_pwd_plain = None
        try:
            default_smtp_port = int(saved_smtp_port) if saved_smtp_port is not None else 465
        except Exception:
            default_smtp_port = 465
        imap_hosts = sorted(self._imap_hosts.keys()) if getattr(self, "_imap_hosts", None) else []
        smtp_suggestions: list[str] = []
        for h in imap_hosts:
            h_low = h.lower()
            if h_low.startswith("imap."):
                smtp_suggestions.append("smtp." + h[5:])
            else:
                smtp_suggestions.append(h)
        if not saved_smtp_host:
            current_imap_host = (self.entry_host.get() or "").strip()
            if current_imap_host:
                h_low = current_imap_host.lower()
                if h_low.startswith("imap."):
                    saved_smtp_host = "smtp." + current_imap_host[5:]
                else:
                    saved_smtp_host = current_imap_host
        host_values = sorted(set(smtp_suggestions + ([saved_smtp_host] if saved_smtp_host else [])))
        combo_smtp_host["values"] = host_values
        if saved_smtp_host:
            smtp_host_var.set(saved_smtp_host)
        elif host_values:
            smtp_host_var.set(host_values[0])
        smtp_port_var.set(str(default_smtp_port))
        self._attach_context_menu(entry_from)
        self._attach_context_menu(entry_to)
        self._attach_context_menu(combo_smtp_host)
        self._attach_context_menu(entry_smtp_port)

        def send_now():
            from_email = (from_var.get() or "").strip()
            to_email = (to_var.get() or "").strip()
            smtp_host = (smtp_host_var.get() or "").strip()
            smtp_port_text = (smtp_port_var.get() or "").strip()
            try:
                sel = list(lb.curselection())
                chosen = [lb.get(i) for i in sel] if sel else []
            except Exception:
                chosen = []
            if not from_email:
                messagebox.showerror("Ошибка", "Укажи e-mail отправителя (From).", parent=dlg)
                return
            if not to_email:
                messagebox.showerror("Ошибка", "Укажи e-mail получателя (To).", parent=dlg)
                return
            if not chosen:
                messagebox.showerror("Ошибка", "Выбери хотя бы один файл для отправки.", parent=dlg)
                return
            if not smtp_host:
                messagebox.showerror("Ошибка", "Укажи SMTP хост.", parent=dlg)
                return
            if not smtp_port_text:
                messagebox.showerror("Ошибка", "Укажи SMTP порт.", parent=dlg)
                return
            try:
                smtp_port_int = int(smtp_port_text)
            except Exception:
                messagebox.showerror("Ошибка", "SMTP порт должен быть числом.", parent=dlg)
                return
            saved_password: str | None = None
            try:
                prev_from = (settings.get("last_report_from") or "").strip()
                prev_to = (settings.get("last_report_to") or "").strip()
                prev_host = (settings.get("smtp_host") or "").strip()
                prev_port = int(settings.get("smtp_port") or 0)
            except Exception:
                prev_from = prev_to = prev_host = ""
                prev_port = 0
            if (
                last_pwd_plain
                and from_email == prev_from
                and to_email == prev_to
                and smtp_host == prev_host
                and smtp_port_int == prev_port
            ):
                saved_password = last_pwd_plain
            pwd = self._ask_password(
                "SMTP пароль",
                f"Пароль для {from_email}:",
                default_value=saved_password,
            )
            if pwd is None or not pwd.strip():
                return
            settings["smtp_from"] = from_email
            settings["smtp_to"] = to_email
            settings["smtp_host"] = smtp_host
            settings["smtp_port"] = smtp_port_int
            settings["last_report_from"] = from_email
            settings["last_report_to"] = to_email
            settings["last_report_password_b64"] = base64.b64encode(
                pwd.encode("utf-8")
            ).decode("ascii")
            self._save_mailsettings(settings)
            try:
                self._send_email_with_attachments(
                    smtp_host=smtp_host,
                    smtp_port=smtp_port_int,
                    from_email=from_email,
                    password=pwd,
                    to_email=to_email,
                    subject="Отчёты ",
                    body="Во вложении отчёты",
                    attachments=chosen,
                )
                messagebox.showinfo(
                    "ОК",
                    "Письмо отправлено.\n\nФайлы:\n" + "\n".join(chosen),
                    parent=dlg,
                )
                dlg.destroy()
            except Exception as e:
                logging.exception("Ошибка отправки письма")
                messagebox.showerror("Ошибка", f"Не удалось отправить письмо:\n{e}", parent=dlg)
        btns = ttk.Frame(frm)
        btns.grid(row=11, column=0, sticky="e")
        ttk.Button(btns, text="Отмена", command=dlg.destroy).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns, text="Отправить", command=send_now).grid(row=0, column=1)
        entry_from.focus_set()

    def _send_email_with_attachments(
        self,
        smtp_host: str,
        smtp_port: int,
        from_email: str,
        password: str,
        to_email: str,
        subject: str,
        body: str,
        attachments: list[str],
    ) -> None:
        msg = EmailMessage()
        msg["From"] = from_email
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.set_content(body or "")

        for path in attachments:
            if not path or not os.path.exists(path):
                continue
            ctype, encoding = mimetypes.guess_type(path)
            if ctype is None or encoding is not None:
                ctype = "application/octet-stream"
            maintype, subtype = ctype.split("/", 1)
            with open(path, "rb") as f:
                data = f.read()
            msg.add_attachment(
                data,
                maintype=maintype,
                subtype=subtype,
                filename=os.path.basename(path),
            )

        with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30) as s:
            s.login(from_email, password)
            s.send_message(msg)

    def _collect_descendants(self, boxes, root_encoded: str):
        delim_map = {}
        all_delims = set()
        for flags_part, decoded_name, encoded_name, delim in boxes:
            if not encoded_name:
                continue
            d = delim if delim and str(delim).upper() != "NIL" else None
            if d:
                delim_map[encoded_name] = d
                all_delims.add(d)
        root_delim = delim_map.get(root_encoded)
        delims = {root_delim} if root_delim else set(all_delims)
        if not delims:
            delims = {"/", "."}
        res = []
        for flags_part, decoded_name, encoded_name, delim in boxes:
            if self._should_skip_folder(flags_part, decoded_name):
                continue
            if encoded_name == root_encoded:
                res.append((decoded_name, encoded_name))
                continue
            for d in delims:
                if encoded_name.startswith(root_encoded + d):
                    res.append((decoded_name, encoded_name))
                    break
        return res

    def _load_imap_hosts(self) -> None:
        self._imap_hosts.clear()
        if not os.path.exists(self._imap_hosts_path):
            return
        try:
            with open(self._imap_hosts_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if ":" not in line:
                        continue
                    host, port_str = line.split(":", 1)
                    host = host.strip()
                    port_str = port_str.strip()
                    if not host or not port_str:
                        continue
                    try:
                        port = int(port_str)
                    except ValueError:
                        continue
                    self._imap_hosts[host] = port
        except Exception as e:
            print(f"Ошибка загрузки IMAP хостов: {e}")
        if not self._imap_hosts:
            self._imap_hosts["imap.yandex.ru"] = 993

    def _on_host_selected(self, event=None) -> None:
        host = self.entry_host.get().strip()
        if host in self._imap_hosts:
            port = self._imap_hosts[host]
            self.entry_port.delete(0, tk.END)
            self.entry_port.insert(0, str(port))

    def _on_date_change(self, var: tk.StringVar, which: str, entry: ttk.Entry) -> None:
        if self._date_guard:
            return
        self._date_guard = True
        try:
            raw = var.get()
            digits = "".join(ch for ch in raw if ch.isdigit())
            digits = digits[:8]
            year = digits[:4] if len(digits) >= 4 else digits
            month = digits[4:6] if len(digits) > 4 else ""
            day = digits[6:8] if len(digits) > 6 else ""
            parts = []
            if year:
                parts.append(year)
            if month:
                parts.append(month)
            if day:
                parts.append(day)
            value = "-".join(parts) if parts else ""
            if value != raw:
                var.set(value)

            def move_cursor() -> None:
                entry.icursor(tk.END)
            entry.after_idle(move_cursor)
        finally:
            self._date_guard = False

    def _attach_context_menu(self, entry: tk.Widget) -> None:
        menu = tk.Menu(entry, tearoff=0)
        menu.add_command(label="Копировать", command=lambda e=entry: e.event_generate("<<Copy>>"))
        menu.add_command(label="Вставить", command=lambda e=entry: e.event_generate("<<Paste>>"))
        menu.add_command(label="Вырезать", command=lambda e=entry: e.event_generate("<<Cut>>"))
        menu.add_separator()
        if isinstance(entry, (tk.Entry, ttk.Entry, ttk.Combobox)):
            menu.add_command(
                label="Выделить всё",
                command=lambda e=entry: (e.select_range(0, "end"), e.icursor("end"))
            )

        def popup(event, m=menu) -> None:
            m.tk_popup(event.x_root, event.y_root)
            m.grab_release()
        entry.bind("<Button-3>", popup)

    def _on_login_selected(self, event=None) -> None:
        return

    def _ask_password(self, title: str, prompt: str, default_value: str | None = None) -> str | None:
        dlg = tk.Toplevel(self)
        dlg.transient(self)
        dlg.title(title or "Пароль")
        dlg.resizable(False, False)
        dlg.grab_set()
        frm = ttk.Frame(dlg, padding=10)
        frm.pack(fill="both", expand=True)
        lbl = ttk.Label(frm, text=prompt)
        lbl.grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))
        pwd_var = tk.StringVar()
        if default_value:
            pwd_var.set(default_value)
        entry = ttk.Entry(frm, show="*", textvariable=pwd_var, width=40)
        entry.grid(row=1, column=0, columnspan=2, sticky="we", pady=(0, 6))
        entry.focus_set()
        entry.icursor(tk.END)
        result: dict[str, str | None] = {"value": None}

        def on_ok(event=None):
            result["value"] = pwd_var.get()
            dlg.destroy()

        def on_cancel(event=None):
            result["value"] = None
            dlg.destroy()

        btn_ok = ttk.Button(frm, text="OK", command=on_ok)
        btn_ok.grid(row=2, column=0, sticky="e", padx=(0, 6))
        btn_cancel = ttk.Button(frm, text="Отмена", command=on_cancel)
        btn_cancel.grid(row=2, column=1, sticky="w")
        dlg.bind("<Return>", on_ok)
        dlg.bind("<Escape>", on_cancel)
        self.update_idletasks()
        dlg.update_idletasks()
        try:
            px = self.winfo_rootx()
            py = self.winfo_rooty()
            pw = self.winfo_width()
            ph = self.winfo_height()
            dw = dlg.winfo_reqwidth()
            dh = dlg.winfo_reqheight()
            dlg.geometry(f"+{px + (pw - dw) // 2}+{py + (ph - dh) // 2}")
        except Exception:
            pass
        dlg.wait_window()
        return result["value"]

    def _add_account(self) -> None:
        login = simpledialog.askstring("Добавить аккаунт", "Логин (email):", parent=self)
        if not login:
            return
        login = login.strip()
        if not login:
            return
        pwd = self._ask_password("Добавить аккаунт", "Пароль:")
        if pwd is None:
            return
        pwd = pwd.strip()
        if not pwd:
            messagebox.showwarning("Внимание", "Пароль пустой — запись не добавлена.", parent=self)
            return
        if login in self._accounts:
            overwrite = messagebox.askyesno(
                "Аккаунт уже есть",
                f"Аккаунт {login} уже существует. Перезаписать пароль?",
                parent=self
            )
            if not overwrite:
                return
        try:
            updated = False
            if os.path.exists(self._accounts_path):
                try:
                    with open(self._accounts_path, "r", encoding="utf-8") as f:
                        lines = f.readlines()
                except Exception:
                    lines = None
                if lines is not None:
                    out_lines = []
                    for line in lines:
                        if line.strip().startswith("#"):
                            out_lines.append(line)
                            continue
                        if ":" not in line:
                            out_lines.append(line)
                            continue
                        lg, _ = line.split(":", 1)
                        if lg.strip() == login:
                            out_lines.append(f"{login}:{pwd}\n")
                            updated = True
                        else:
                            out_lines.append(line)
                    try:
                        with open(self._accounts_path, "w", encoding="utf-8") as f:
                            f.writelines(out_lines)
                    except Exception:
                        with open(self._accounts_path, "a", encoding="utf-8") as f:
                            f.write(f"{login}:{pwd}\n")
                            updated = True
            if not updated:
                with open(self._accounts_path, "a", encoding="utf-8") as f:
                    f.write(f"{login}:{pwd}\n")
            self._accounts[login] = pwd
            self._refresh_accounts_listbox()
            try:
                vals = sorted(self._accounts.keys())
                idx = vals.index(login) if login in vals else None
                if idx is not None and hasattr(self, "accounts_listbox"):
                    self.accounts_listbox.selection_clear(0, tk.END)
                    self.accounts_listbox.selection_set(idx)
                    self.accounts_listbox.see(idx)
            except Exception:
                pass
            messagebox.showinfo("ОК", f"Аккаунт {login} сохранён.", parent=self)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить аккаунт:\n{e}", parent=self)


    def _delete_selected_accounts(self, event=None) -> str:
        if getattr(self, "_busy", False):
            messagebox.showwarning("Внимание", "Нельзя изменять список аккаунтов во время расчёта.", parent=self)
            return "break"
        if not hasattr(self, "accounts_listbox"):
            return "break"
        try:
            sel = list(self.accounts_listbox.curselection())
        except Exception:
            sel = []
        if not sel:
            return "break"
        accounts: list[str] = []
        for i in sel:
            try:
                v = (self.accounts_listbox.get(i) or "").strip()
            except Exception:
                v = ""
            if v:
                accounts.append(v)
        seen: set[str] = set()
        uniq_accounts: list[str] = []
        for a in accounts:
            if a in seen:
                continue
            seen.add(a)
            uniq_accounts.append(a)
        accounts = uniq_accounts
        if not accounts:
            return "break"
        if len(accounts) == 1:
            prompt = (
                f"Удалить аккаунт?\n\n{accounts[0]}\n\n"
                "Он будет удалён из списка и из файла accounts.csv."
            )
        else:
            shown = "\n".join(accounts[:10])
            tail = "" if len(accounts) <= 10 else f"\n… и ещё {len(accounts) - 10}"
            prompt = (
                f"Удалить выбранные аккаунты? ({len(accounts)} шт.)?\n\n"
                f"{shown}{tail}\n\n"
                "Они будут удалены из списка и из файла accounts.csv."
            )
        if not messagebox.askyesno("Удаление аккаунтов", prompt, parent=self):
            return "break"
        try:
            if os.path.exists(self._accounts_path):
                with open(self._accounts_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                out_lines: list[str] = []
                to_remove = set(accounts)
                for line in lines:
                    s = line.strip()
                    if not s or s.startswith("#") or ":" not in s:
                        out_lines.append(line)
                        continue
                    lg, _ = line.split(":", 1)
                    if lg.strip() in to_remove:
                        continue
                    out_lines.append(line)
                with open(self._accounts_path, "w", encoding="utf-8") as f:
                    f.writelines(out_lines)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить аккаунты из accounts.csv:\n{e}", parent=self)
            return "break"
        for a in accounts:
            self._accounts.pop(a, None)

        first_old_index = sel[0] if sel else 0
        self._refresh_accounts_listbox()
        try:
            vals = sorted(self._accounts.keys())
            if vals:
                idx = min(first_old_index, len(vals) - 1)
                self.accounts_listbox.selection_clear(0, tk.END)
                self.accounts_listbox.selection_set(idx)
                self.accounts_listbox.see(idx)
        except Exception:
            pass
        return "break"

    def _build_ui(self) -> None:
        padding = {"padx": 5, "pady": 3}
        main_frame = ttk.Frame(self)
        main_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        row = 0
        ttk.Label(main_frame, text="IMAP хост:").grid(row=row, column=0, sticky="w", **padding)
        host_values = sorted(self._imap_hosts.keys()) if self._imap_hosts else []
        self.entry_host = ttk.Combobox(main_frame, width=28, values=host_values)
        self.entry_host.grid(row=row, column=1, **padding)
        if host_values:
            self.entry_host.set(host_values[0])
        else:
            self.entry_host.set("imap.yandex.ru")
        self.entry_host.bind("<<ComboboxSelected>>", self._on_host_selected)
        self._attach_context_menu(self.entry_host)
        row += 1
        ttk.Label(main_frame, text="IMAP порт:").grid(row=row, column=0, sticky="w", **padding)
        self.entry_port = ttk.Entry(main_frame, width=10)
        self.entry_port.grid(row=row, column=1, sticky="w", **padding)
        self.entry_port.insert(0, "993")
        self._attach_context_menu(self.entry_port)
        row += 1
        ttk.Label(main_frame, text="Ящики для анализа:").grid(row=row, column=0, sticky="w", **padding)
        btn_add_account = ttk.Button(main_frame, text="+", width=3, command=self._add_account)
        btn_add_account.grid(row=row, column=2, sticky="w", padx=(0, 5))
        row += 1
        self.accounts_listbox = tk.Listbox(main_frame, height=7, selectmode=tk.EXTENDED, exportselection=False)
        self.accounts_listbox.grid(row=row, column=1, sticky="we", **padding)
        self.accounts_listbox.bind("<Delete>", self._delete_selected_accounts)

        try:
            self.accounts_listbox.delete(0, tk.END)
            for acc in sorted(self._accounts.keys()):
                self.accounts_listbox.insert(tk.END, acc)
        except Exception:
            pass
        row += 1
        ttk.Label(main_frame,
          text="Ctrl — отдельные, Shift — диапазон",
        ).grid(row=row, column=0, columnspan=2, sticky="w", **padding)
        row += 1
        ttk.Label(main_frame, text="Папка входящих:").grid(row=row, column=0, sticky="w", **padding)
        self.entry_inbox = ttk.Entry(main_frame, width=30)
        self.entry_inbox.grid(row=row, column=1, **padding)
        self.entry_inbox.insert(0, "Входящие")
        self._attach_context_menu(self.entry_inbox)
        row += 1
        ttk.Label(main_frame, text="Папка исходящих:").grid(row=row, column=0, sticky="w", **padding)
        self.entry_sent = ttk.Entry(main_frame, width=30)
        self.entry_sent.grid(row=row, column=1, **padding)
        self.entry_sent.insert(0, "Отправленные")
        self._attach_context_menu(self.entry_sent)
        row += 1
        ttk.Label(main_frame, text="Фильтр домена для исходящих:").grid(row=row, column=0, sticky="w", **padding)
        self.entry_sent_domain = ttk.Entry(main_frame, width=30)
        self.entry_sent_domain.grid(row=row, column=1, **padding)
        self._attach_context_menu(self.entry_sent_domain)
        row += 1
        ttk.Label(main_frame, text="Фильтр домена для входящих:").grid(row=row, column=0, sticky="w", **padding)
        self.entry_inbox_domain = ttk.Entry(main_frame, width=30)
        self.entry_inbox_domain.grid(row=row, column=1, **padding)
        self._attach_context_menu(self.entry_inbox_domain)
        row += 1
        detect_btn = ttk.Button(main_frame, text="Определить папки", command=self.on_detect_folders)
        detect_btn.grid(row=row, column=0, columnspan=2, pady=(5, 5))
        row += 1
        ttk.Label(main_frame, text="Дата с (YYYY-MM-DD):").grid(row=row, column=0, sticky="w", **padding)
        self.from_date_var = tk.StringVar()
        self.entry_from_date = ttk.Entry(main_frame, width=30, textvariable=self.from_date_var)
        self.entry_from_date.grid(row=row, column=1, **padding)
        self.from_date_var.trace_add("write", lambda *args: self._on_date_change(self.from_date_var, "from", self.entry_from_date))
        self._attach_context_menu(self.entry_from_date)
        row += 1
        ttk.Label(main_frame, text="Дата по (YYYY-MM-DD):").grid(row=row, column=0, sticky="w", **padding)
        self.to_date_var = tk.StringVar()
        self.entry_to_date = ttk.Entry(main_frame, width=30, textvariable=self.to_date_var)
        self.entry_to_date.grid(row=row, column=1, **padding)
        self.to_date_var.trace_add("write", lambda *args: self._on_date_change(self.to_date_var, "to", self.entry_to_date))
        self._attach_context_menu(self.entry_to_date)
        row += 1
        self.btn_calc = ttk.Button(main_frame, text="Посчитать", command=self.on_calculate)
        self.btn_calc.grid(row=row, column=0, columnspan=2, pady=(10, 5))
        row += 1
        self.progress = ttk.Progressbar(main_frame, mode="determinate", length=260)
        self.progress.grid(row=row, column=0, columnspan=2, sticky="we", pady=(0, 6))
        row += 1
        self.label_status = ttk.Label(main_frame, text="Статус: готово")
        self.label_status.grid(row=row, column=0, columnspan=2, sticky="w", **padding)
        row += 1
        self.btn_cancel = ttk.Button(main_frame, text="Отмена", command=self.on_cancel)
        self.btn_cancel.grid(row=row, column=0, columnspan=2, pady=(0, 8))
        self.btn_cancel.config(state="disabled")
        row += 1
        self.label_result_inbox = ttk.Label(main_frame, text="Всего писем: -")
        self.label_result_inbox.grid(row=row, column=0, columnspan=2, sticky="w", **padding)
        row += 1
        self.label_result_sent = ttk.Label(main_frame, text="Проверено папок: -")
        self.label_result_sent.grid(row=row, column=0, columnspan=2, sticky="w", **padding)
        row += 1
        self.btn_send_reports = ttk.Button(main_frame, text="Отправить отчёты", command=self.on_send_reports)
        self.btn_send_reports.grid(row=row, column=0, columnspan=2, pady=(8, 0))
        self.btn_send_reports.config(state="disabled")

    def _load_config(self) -> None:
        if not os.path.exists(self._config_path):
            return
        try:
            with open(self._config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception:
            return
        host = cfg.get("host")
        port = cfg.get("port")
        inbox = cfg.get("inbox")
        sent = cfg.get("sent")
        sent_domain = cfg.get("sent_domain")
        inbox_domain = cfg.get("inbox_domain")
        from_date = cfg.get("from_date")
        to_date = cfg.get("to_date")
        if host:
            self.entry_host.delete(0, tk.END)
            self.entry_host.insert(0, host)
        if port:
            self.entry_port.delete(0, tk.END)
            self.entry_port.insert(0, str(port))
        if inbox:
            self.entry_inbox.delete(0, tk.END)
            self.entry_inbox.insert(0, inbox)
        if sent:
            self.entry_sent.delete(0, tk.END)
            self.entry_sent.insert(0, sent)
        if sent_domain:
            self.entry_sent_domain.delete(0, tk.END)
            self.entry_sent_domain.insert(0, sent_domain)
        if inbox_domain:
            self.entry_inbox_domain.delete(0, tk.END)
            self.entry_inbox_domain.insert(0, inbox_domain)
        if from_date:
            self.from_date_var.set(from_date)
        if to_date:
            self.to_date_var.set(to_date)

    def _save_config(self) -> None:
        cfg = self._load_mailsettings()
        if not isinstance(cfg, dict):
            cfg = {}
        cfg.update(
            {
                "host": self.entry_host.get().strip(),
                "port": self.entry_port.get().strip(),
                "inbox": self.entry_inbox.get().strip(),
                "sent": self.entry_sent.get().strip(),
                "sent_domain": self.entry_sent_domain.get().strip(),
                "inbox_domain": self.entry_inbox_domain.get().strip(),
                "from_date": self.from_date_var.get().strip(),
                "to_date": self.to_date_var.get().strip(),
            }
        )
        try:
            self._save_mailsettings(cfg)
        except Exception as e:
            print(f"Ошибка сохранения конфига: {e}")

    def _iter_mailboxes(self, mail):
        boxes = []

        def _try_list():
            for args in (None, ('', '*'), ('""', '*')):
                try:
                    if args is None:
                        typ, data = mail.list()
                    else:
                        typ, data = mail.list(*args)
                    if typ == "OK" and data:
                        return data
                except Exception:
                    continue
            return []
        data = _try_list()
        if not data:
            return boxes
        rx = re.compile(r'^\((?P<flags>[^)]*)\)\s+(?P<delim>NIL|"[^"]*")\s+(?P<name>.+)$', re.IGNORECASE)
        for raw in data:
            if not raw:
                continue
            line = raw.decode("utf-8", errors="replace") if isinstance(raw, (bytes, bytearray)) else str(raw)
            line = line.strip()
            m = rx.match(line)
            if not m:
                logging.info("LIST: не удалось разобрать строку: %r", line)
                continue
            flags_part = (m.group("flags") or "").strip()
            delim_tok = (m.group("delim") or "").strip()
            if delim_tok.upper() == "NIL":
                delim = None
            else:
                delim = delim_tok.strip('"')
            name_tok = (m.group("name") or "").strip()
            if len(name_tok) >= 2 and name_tok.startswith('"') and name_tok.endswith('"'):
                encoded_name = name_tok[1:-1]
            else:
                encoded_name = name_tok
            encoded_name = encoded_name.strip()
            if not encoded_name:
                continue
            decoded_name = decode_imap_utf7(encoded_name)
            boxes.append((flags_part, decoded_name, encoded_name, delim))
        return boxes

    def _is_drafts_folder(self, flags_part: str, decoded_name: str) -> bool:
        flags_upper = (flags_part or "").upper()
        name_upper = (decoded_name or "").upper()
        if "\\DRAFTS" in flags_upper:
            return True
        drafts_keywords = [
            "DRAFTS",
            "DRAFT",
            "ЧЕРНОВИКИ",
            "ЧЕРНОВЫЕ",
        ]
        for kw in drafts_keywords:
            if kw in name_upper:
                return True
        return False

    def _last_seg_upper(self, name: str) -> str:
        s = (name or "").strip()
        for sep in ("/", "\\", "."):
            if sep in s:
                s = s.split(sep)[-1]
        return s.strip().upper()

    def _is_sent_folder(self, flags_part: str, decoded_name: str) -> bool:
        flags_upper = (flags_part or "").upper()
        last = self._last_seg_upper(decoded_name)
        sent_cfg = self.entry_sent.get().strip()
        sent_cfg_upper = sent_cfg.upper() if sent_cfg else ""
        if "\\SENT" in flags_upper:
            return True
        if sent_cfg_upper and last == sent_cfg_upper:
            return True
        return last in {
            "SENT",
            "SENT ITEMS",
            "ОТПРАВЛЕННЫЕ",
            "ОТПРАВЛЕННЫЕ ПИСЬМА",
        }

    def _is_inbox_folder(self, flags_part: str, decoded_name: str) -> bool:
        flags_upper = (flags_part or "").upper()
        name_upper = (decoded_name or "").upper()
        inbox_cfg = self.entry_inbox.get().strip()
        inbox_cfg_upper = inbox_cfg.upper() if inbox_cfg else ""
        if "\\INBOX" in flags_upper:
            return True
        if inbox_cfg_upper and name_upper == inbox_cfg_upper:
            return True
        if name_upper == "INBOX":
            return True
        if "ВХОДЯЩИЕ" in name_upper:
            return True
        return False

    def _is_trash_folder(self, flags_part: str, decoded_name: str) -> bool:
        flags_upper = (flags_part or "").upper()
        last = self._last_seg_upper(decoded_name)
        if "\\TRASH" in flags_upper:
            return True
        return last in {
            "DELETED ITEMS",
            "DELETED MESSAGES",
            "TRASH",
            "BIN",
            "КОРЗИНА",
            "УДАЛЕННЫЕ",
            "УДАЛЁННЫЕ",
        }

    def _is_spam_folder(self, flags_part: str, decoded_name: str) -> bool:
        flags_upper = (flags_part or "").upper()
        name_upper = (decoded_name or "").upper()
        if "\\JUNK" in flags_upper or "\\SPAM" in flags_upper:
            return True
        spam_keywords = [
            "SPAM",
            "JUNK",
            "СПАМ",
            "НЕЖЕЛАТЕЛЬНЫЕ",
        ]
        for kw in spam_keywords:
            if kw in name_upper:
                return True
        return False

    def _should_skip_folder(self, flags_part: str, decoded_name: str) -> bool:
        if decoded_name in ("|", "|:") or (decoded_name or "").startswith("|:"):
            return True
        if self._is_drafts_folder(flags_part, decoded_name):
            return True
        if self._is_spam_folder(flags_part, decoded_name):
            return True
        if self._is_trash_folder(flags_part, decoded_name):
            return True
        return False


    def _select_main_folders(self, boxes: list[tuple[str, str, str, str]]) -> tuple[str | None, str | None]:
        inbox_cfg = (self.entry_inbox.get() or "").strip()
        sent_cfg = (self.entry_sent.get() or "").strip()
        inbox_cfg_up = inbox_cfg.upper()
        sent_cfg_up = sent_cfg.upper()
        inbox_by_cfg: str | None = None
        inbox_by_flag: str | None = None
        inbox_first: str | None = None
        sent_by_cfg: str | None = None
        sent_by_flag: str | None = None
        sent_first: str | None = None
        for flags_part, decoded_name, encoded_name, delim in boxes:
            name_up = (decoded_name or "").upper()
            flags_up = (flags_part or "").upper()
            if self._is_inbox_folder(flags_part, decoded_name):
                if inbox_first is None:
                    inbox_first = encoded_name
                if inbox_cfg_up and name_up == inbox_cfg_up:
                    inbox_by_cfg = encoded_name
                if "\\INBOX" in flags_up:
                    inbox_by_flag = encoded_name
            if self._is_sent_folder(flags_part, decoded_name):
                if sent_first is None:
                    sent_first = encoded_name
                if sent_cfg_up and self._last_seg_upper(decoded_name) == sent_cfg_up:
                    sent_by_cfg = encoded_name
                if "\\SENT" in flags_up:
                    sent_by_flag = encoded_name
        main_inbox = inbox_by_cfg or inbox_by_flag or inbox_first
        main_sent = sent_by_cfg or sent_by_flag or sent_first
        return main_inbox, main_sent

    def on_detect_folders(self) -> None:
        host = self.entry_host.get().strip()
        port_str = self.entry_port.get().strip()
        selected = self._get_selected_accounts()
        if not selected:
            messagebox.showerror("Ошибка", "Выбери один ящик для определения папок.", parent=self)
            return
        if len(selected) > 1:
            messagebox.showerror("Ошибка", "Определение папок выполняется только для одного ящика. Выбери один.", parent=self)
            return
        username = selected[0].strip()
        password = (self._accounts.get(username) or "").strip()
        if not host:
            messagebox.showerror("Ошибка", "Укажи IMAP хост.", parent=self)
            return
        if not username:
            messagebox.showerror("Ошибка", "Укажи логин (e-mail).", parent=self)
            return
        if not password:
            messagebox.showerror("Ошибка", f"Для {username} нет пароля в accounts.csv.", parent=self)
            return
        try:
            port = int(port_str)
        except ValueError:
            messagebox.showerror("Ошибка", "Порт должен быть числом.", parent=self)
            return
        self._set_busy(True, status="определяю папки")
        try:
            self.progress["value"] = 0
        except Exception:
            pass

        def worker():
            mail = None
            try:
                mail = connect_imap(host, port, username, password, timeout=30.0)
                boxes = self._iter_mailboxes(mail)
                main_inbox, main_sent = self._select_main_folders(boxes)
                enc_to_dec = {en: dn for _fl, dn, en, _dl in boxes if en}

                def last_segment(name: str) -> str:
                    s = (name or "").strip()
                    for sep in ("/", "\\", "."):
                        if sep in s:
                            s = s.split(sep)[-1]
                    return s.strip()
                
                inbox_dec = enc_to_dec.get(main_inbox, "") if main_inbox else ""
                sent_dec = enc_to_dec.get(main_sent, "") if main_sent else ""
                inbox_ui = "Входящие" if (inbox_dec or "").upper() == "INBOX" else (last_segment(inbox_dec) or "Входящие")
                sent_ui = last_segment(sent_dec) or "Отправленные"

                def finish_ok():
                    self.entry_inbox.delete(0, tk.END)
                    self.entry_inbox.insert(0, inbox_ui)
                    self.entry_sent.delete(0, tk.END)
                    self.entry_sent.insert(0, sent_ui)
                    self._set_busy(False, status="готово")
                    messagebox.showinfo("ОК", "Папки определены и подставлены в поля.", parent=self)

                self.after(0, finish_ok)

            except Exception as e:
                logging.exception("Ошибка определения папок")
                self.after(0, lambda err=str(e): (
                self._set_busy(False, status="готово"),
                messagebox.showerror("Ошибка", f"Не удалось определить папки:\n{err}", parent=self),
            ))

            finally:
                try:
                    if mail is not None:
                        try:
                            mail.close()
                        except Exception:
                            pass
                        try:
                            mail.logout()
                        except Exception:
                            pass
                except Exception:
                    pass

        self._start_worker(worker, ())

    def _display_folder_name(self, dn: str) -> str:
        s = (dn or "").strip()
        up = s.upper()
        if up == "INBOX":
            return "Входящие"
        for sep in ("/", "\\", "."):
            pref = "INBOX" + sep
            if up.startswith(pref):
                return "Входящие" + sep + s[len(pref):]
        return s

    def _set_busy(self, busy: bool, status: str | None = None) -> None:
        self._busy = busy
        self.btn_calc.config(state=("disabled" if busy else "normal"))
        self.btn_cancel.config(state=("normal" if busy else "disabled"))
        if status is not None:
            self.label_status.config(text=f"Статус: {status}")
        if not busy:
            try:
                self.progress["value"] = 0
            except Exception:
                pass

    def _update_progress(self, current: int, total: int, status: str | None = None) -> None:
        try:
            total = max(1, int(total))
            current = max(0, int(current))
            pct = min(100.0, (current / total) * 100.0)
            self.progress["value"] = pct
        except Exception:
            pass
        if status is not None:
            self.label_status.config(text=f"Статус: {status}")
        self.label_result_sent.config(text=f"Проверено папок: {current}/{total}")
        self.update_idletasks()

    def on_cancel(self) -> None:
        if not self._busy:
            return
        self._cancel_event.set()
        self.label_status.config(text="Статус: останавливаю расчёт")

    def _start_worker(self, target, args: tuple) -> None:
        if self._worker_thread and self._worker_thread.is_alive():
            return
        self._cancel_event.clear()
        self._worker_thread = threading.Thread(target=target, args=args, daemon=True)
        self._worker_thread.start()


    def _validate_common_params(self):
        host = (self.entry_host.get() or "").strip()
        port_str = (self.entry_port.get() or "").strip()

        inbox_folder = (self.entry_inbox.get() or "").strip()
        sent_folder = (self.entry_sent.get() or "").strip()

        from_date = (self.from_date_var.get() or "").strip()
        to_date = (self.to_date_var.get() or "").strip()

        sent_domain = (self.entry_sent_domain.get() or "").strip()
        inbox_domain = (self.entry_inbox_domain.get() or "").strip()

        if not host:
            raise ValueError("Укажи IMAP хост.")
        try:
            port = int(port_str)
        except Exception:
            raise ValueError("Порт должен быть числом.")
        if not inbox_folder:
            raise ValueError("Укажи название папки входящих (например: Входящие).")
        if not sent_folder:
            raise ValueError("Укажи название папки исходящих (например: Отправленные).")
        if not from_date or not to_date:
            raise ValueError("Для работы нужно задать обе даты: 'Дата с' и 'Дата по'.")
        try:
            d1 = datetime.strptime(from_date, "%Y-%m-%d")
            d2 = datetime.strptime(to_date, "%Y-%m-%d")
        except Exception:
            raise ValueError("Дата должна быть в формате YYYY-MM-DD.")
        if d1 > d2:
            raise ValueError("Дата 'с' не может быть позже даты 'по'.")
        return host, port, inbox_folder, sent_folder, from_date, to_date, sent_domain, inbox_domain

    def on_calculate(self) -> None:
        try:
            (
                host,
                port,
                inbox_folder,
                sent_folder,
                from_date,
                to_date,
                sent_domain,
                inbox_domain,
            ) = self._validate_common_params()
        except ValueError as e:
            messagebox.showerror("Ошибка", str(e), parent=self)
            return

        accounts = self._get_selected_accounts()
        if not accounts:
            messagebox.showerror("Ошибка", "Выбери один или несколько ящиков для анализа.", parent=self)
            return

        sent_domain_norm = sent_domain.strip() if sent_domain else None
        inbox_domain_norm = inbox_domain.strip() if inbox_domain else None

        inbox_cfg_ui = (self.entry_inbox.get() or "").strip()
        sent_cfg_ui = (self.entry_sent.get() or "").strip()

        self.label_result_inbox.config(text="Всего писем: считаем...")
        self.label_result_sent.config(text="Проверено папок: готовлю список…")
        try:
            self.progress["value"] = 0
        except Exception:
            pass
        self.label_result_inbox.config(text="Всего писем: считаем.")
        self.label_result_sent.config(text="Проверено папок: готовлю список…")
        try:
            self.progress["value"] = 0
        except Exception:
            pass
        self._last_report_files = []
        try:
            self.btn_send_reports.config(state="disabled")
        except Exception:
            pass
        self._set_busy(True, status="подключаюсь к почте")
        if len(accounts) == 1:
            username = accounts[0]
            password = (self._accounts.get(username) or "").strip()
            if not password:
                self._set_busy(False, status="готово")
                messagebox.showerror("Ошибка", f"Для {username} нет пароля в accounts.csv.", parent=self)
                return
            self._start_worker(
                self._calculate_worker,
                (host, port, username, password, from_date, to_date,
                 sent_domain_norm, inbox_domain_norm, inbox_cfg_ui, sent_cfg_ui),
            )
        else:
            missing = [u for u in accounts if not (self._accounts.get(u) or "").strip()]
            if missing:
                self._set_busy(False, status="готово")
                messagebox.showerror("Ошибка", "Для выбранных ящиков нет пароля в accounts.csv:\n" + "\n".join(missing), parent=self)
                return
            self._start_worker(
                self._calculate_multi_worker,
                (host, port, accounts, from_date, to_date,
                 sent_domain_norm, inbox_domain_norm, inbox_cfg_ui, sent_cfg_ui),
            )
    def _calculate_core(
            self,
            host: str,
            port: int,
            username: str,
            password: str,
            from_date: str,
            to_date: str,
            sent_domain_norm: str | None,
            inbox_domain_norm: str | None,
            inbox_cfg_ui: str,
            sent_cfg_ui: str,
        ) -> dict:
        mail = None
        cache_path = os.path.join(BASE_DIR, "mid_cache.json")
        cache_obj = _load_mid_cache(cache_path)
        folders_cache = cache_obj.get("folders")
        if not isinstance(folders_cache, dict):
            folders_cache = {}
            cache_obj["folders"] = folders_cache
        try:
            mail = connect_imap(host, port, username, password, timeout=30.0)
            boxes = self._iter_mailboxes(mail)
            main_inbox, main_sent = self._select_main_folders(boxes)

            enc_to_flags: dict[str, str] = {}
            for flags, dn, en, _dl in boxes:
                if en:
                    enc_to_flags[en] = flags or ""

            def check_cancel() -> None:
                if self._cancel_event.is_set():
                    raise RuntimeError("Отменено пользователем.")

            inbox_list: list[tuple[str, str]] = []
            sent_list: list[tuple[str, str]] = []
            if main_inbox:
                inbox_list = self._collect_descendants(boxes, main_inbox)
            if main_sent:
                sent_list = self._collect_descendants(boxes, main_sent)

            def _uniq_by_enc(pairs: list[tuple[str, str]]) -> list[tuple[str, str]]:
                seen_enc: set[str] = set()
                out: list[tuple[str, str]] = []
                for dn, en in pairs:
                    if not en or en in seen_enc:
                        continue
                    seen_enc.add(en)
                    out.append((dn, en))
                return out

            inbox_list = _uniq_by_enc(inbox_list)
            sent_list = _uniq_by_enc(sent_list)
            inbox_list = [(dn, en) for dn, en in inbox_list if not self._should_skip_folder(enc_to_flags.get(en, ""), dn)]
            sent_list = [(dn, en) for dn, en in sent_list if not self._should_skip_folder(enc_to_flags.get(en, ""), dn)]
            total_folders = max(1, len(inbox_list) + len(sent_list))
            processed = 0

            def tick(status: str | None = None) -> None:
                self.after(0, lambda: self._update_progress(processed, total_folders, status))
            self.after(0, lambda: self.label_status.config(text=f"Статус: считаю письма ({username})"))
            tick()

            seen_inbox: set[bytes] = set()
            seen_sent: set[bytes] = set()

            def _seen_digest(key: str) -> bytes:
                return hashlib.blake2b(key.encode("utf-8", "ignore"), digest_size=8).digest()

            unique_inbox = 0
            unique_sent = 0

            def build_inbox_criteria(domain: str | None) -> list[str] | None:
                if not domain:
                    return None
                return ["FROM", domain]

            def build_sent_criteria(domain: str | None) -> list[str] | None:
                if not domain:
                    return None
                return ["OR",
                        "HEADER", "To", domain,
                        "OR",
                            "HEADER", "Cc", domain,
                            "HEADER", "Bcc", domain
                    ]

            def _cache_key(enc_folder: str) -> str:
                return f"{host}|{username}|{enc_folder}"

            def ensure_folder_cache(enc_folder: str) -> dict:
                ckey = _cache_key(enc_folder)
                entry = folders_cache.get(ckey)
                if not isinstance(entry, dict):
                    entry = {}
                    folders_cache[ckey] = entry
                mp = entry.get("map")
                if not isinstance(mp, dict):
                    mp = {}
                    entry["map"] = mp
                uidv_now = _get_uidvalidity(mail, enc_folder)
                uidv_old = str(entry.get("uidvalidity") or "")
                if uidv_now:
                    if uidv_old and uidv_old != uidv_now:
                        mp.clear()
                    entry["uidvalidity"] = uidv_now
                return entry

            def process_folder(kind: str, dec_name: str, enc_name: str) -> None:
                nonlocal processed, unique_inbox, unique_sent
                check_cancel()
                try:
                    if kind == "inbox":
                        extra = build_inbox_criteria(inbox_domain_norm)
                        use_sent_date = False
                    else:
                        extra = build_sent_criteria(sent_domain_norm)
                        use_sent_date = False

                    uids_b = search_uids(mail, enc_name, from_date, to_date, extra_criteria=extra, include_deleted=False, use_sent_date=use_sent_date)

                    if kind == "sent" and sent_domain_norm and not uids_b:
                        try:
                            check_cancel()
                            uids_b2 = search_uids(mail, enc_name, from_date, to_date, extra_criteria=extra, include_deleted=False, use_sent_date=True)
                            if uids_b2:
                                uids_b = uids_b2
                        except Exception:
                            pass

                    if not uids_b:
                        return

                    entry = ensure_folder_cache(enc_name)
                    mp = entry.get("map")
                    if not isinstance(mp, dict):
                        mp = {}
                        entry["map"] = mp

                    uids: list[str] = []
                    missing: list[str] = []
                    for ub in uids_b:
                        check_cancel()
                        uid = ub.decode("ascii", "ignore") if isinstance(ub, (bytes, bytearray)) else str(ub)
                        uid = uid.strip()
                        if not uid:
                            continue
                        uids.append(uid)
                        if uid not in mp:
                            missing.append(uid)

                    if missing:
                        check_cancel()
                        missing_b = [x.encode("ascii", "ignore") for x in missing]
                        got = fetch_message_id_map(mail, missing_b, step=800)
                        for uid, key in got.items():
                            if key:
                                mp[uid] = key

                    entry["map"] = mp

                    for uid in uids:
                        check_cancel()
                        key = mp.get(uid) or ("uid:" + uid)
                        dg = _seen_digest(key)
                        if kind == "inbox":
                            if dg in seen_inbox:
                                continue
                            seen_inbox.add(dg)
                            unique_inbox += 1
                        else:
                            if dg in seen_sent:
                                continue
                            seen_sent.add(dg)
                            unique_sent += 1

                finally:
                    processed += 1
                    if kind == "inbox":
                        tick(f"Проверяю папку «{self._display_folder_name(dec_name)}» (входящие) [{username}]")
                    else:
                        tick(f"Проверяю папку «{self._display_folder_name(dec_name)}» (исходящие) [{username}]")

            for dec_name, enc_name in inbox_list:
                process_folder("inbox", dec_name, enc_name)
            for dec_name, enc_name in sent_list:
                process_folder("sent", dec_name, enc_name)

            total_unique = unique_inbox + unique_sent
            label_line = f"Писем за период: {total_unique} (входящие: {unique_inbox}, исходящие: {unique_sent})"

            msg_lines = [
                f"Всего писем за период (входящие + исходящие): {total_unique}",
                f"Входящие (уникальные): {unique_inbox}",
                f"Исходящие (уникальные): {unique_sent}",
                "",
                f"Период: с {from_date} по {to_date}",
                f"Аккаунт: {username}",
                f"Папок проверено всего (входящие + исходящие): {processed}",
            ]
            if inbox_cfg_ui:
                msg_lines.append(f"Папка входящих: {inbox_cfg_ui}")
            if sent_cfg_ui:
                msg_lines.append(f"Папка исходящих: {sent_cfg_ui}")
            msg_lines.append(f"Папок проверено (входящие): {len(inbox_list)}")
            msg_lines.append(f"Папок проверено (исходящие): {len(sent_list)}")

            result = {
                "ok": True,
                "title": "Результат",
                "username": username,
                "from_date": from_date,
                "to_date": to_date,
                "message": "\n".join(msg_lines),
                "label_line": label_line,
                "folders_processed": processed,
                "total_unique": total_unique,
                "unique_inbox": unique_inbox,
                "unique_sent": unique_sent,
                "silent": False,
            }

            _save_mid_cache(cache_path, cache_obj)
            return result

        except Exception as e:
            logging.exception("Ошибка расчёта")
            return {
                "ok": False,
                "username": username,
                "from_date": from_date,
                "to_date": to_date,
                "error": str(e),
                "trace": traceback.format_exc(),
                "silent": False,
            }
        finally:
            try:
                if mail is not None:
                    try:
                        mail.close()
                    except Exception:
                        pass
                    try:
                        mail.logout()
                    except Exception:
                        pass
            except Exception:
                pass

    def _calculate_worker(
            self,
            host: str,
            port: int,
            username: str,
            password: str,
            from_date: str,
            to_date: str,
            sent_domain_norm: str | None,
            inbox_domain_norm: str | None,
            inbox_cfg_ui: str,
            sent_cfg_ui: str,
        ) -> None:
        res = self._calculate_core(host, port, username, password, from_date, to_date, sent_domain_norm, inbox_domain_norm, inbox_cfg_ui, sent_cfg_ui)
        self.after(0, lambda r=res: self._finish_calculation(r))

    def _calculate_multi_worker(
            self,
            host: str,
            port: int,
            usernames: list[str],
            from_date: str,
            to_date: str,
            sent_domain_norm: str | None,
            inbox_domain_norm: str | None,
            inbox_cfg_ui: str,
            sent_cfg_ui: str,
        ) -> None:
        summary_lines: list[str] = []
        ok_count = 0
        err_count = 0
        last_processed = 0

        for i, u in enumerate(usernames, start=1):
            if self._cancel_event.is_set():
                break
            pwd = (self._accounts.get(u) or "").strip()
            if not pwd:
                err_count += 1
                summary_lines.append(f"{i}) {u}: нет пароля в accounts.csv")
                continue

            self.after(0, lambda n=i, t=len(usernames), a=u: self.label_status.config(text=f"Статус: анализ {n}/{t} — {a}"))

            res = self._calculate_core(host, port, u, pwd, from_date, to_date, sent_domain_norm, inbox_domain_norm, inbox_cfg_ui, sent_cfg_ui)
            res["multi_mode"] = True
            res["silent"] = True
            self.after(0, lambda r=res: self._finish_calculation(r))

            if res.get("ok"):
                ok_count += 1
                last_processed = int(res.get("folders_processed", 0) or 0)
                summary_lines.append(f"{i}) {u}: всего {res.get('total_unique', 0)}, входящие {res.get('unique_inbox', 0)}, исходящие {res.get('unique_sent', 0)}")
            else:
                err_count += 1
                summary_lines.append(f"{i}) {u}: ошибка — {res.get('error', 'неизвестная ошибка')}")

        def finish():
            self.label_status.config(text="Статус: готово")
            self.label_result_inbox.config(text=f"Готово. Успешно: {ok_count}, ошибок: {err_count}")
            self.label_result_sent.config(text=f"Проверено папок: {last_processed}")
            try:
                self.progress["value"] = 100
            except Exception:
                pass
            messagebox.showinfo("Результат (несколько ящиков)", "Результаты по ящикам:\n\n" + "\n".join(summary_lines), parent=self)
            self._set_busy(False)
            self._save_config()

        self.after(0, finish)

    def _finish_calculation(self, result: dict) -> None:
        try:
            ok = bool(result.get("ok"))
            if ok:
                self.label_result_inbox.config(
                    text=result.get("label_line", "Всего писем: -")
                )
                self.label_result_sent.config(
                    text=f"Проверено папок: {result.get('folders_processed', 0)}"
                )
                self.label_status.config(text="Статус: готово")
                try:
                    self.progress["value"] = 100
                except Exception:
                    pass
                title = result.get("title", "Результат")
                username = result.get("username")
                from_date = result.get("from_date")
                to_date = result.get("to_date")
                full_msg = result.get("message", "") or ""
                report_txt_path = save_full_report(
                    title=title,
                    message=full_msg,
                    username=username,
                    from_date=from_date,
                    to_date=to_date,
                )
                total_unique = result.get("total_unique")
                unique_inbox = result.get("unique_inbox")
                unique_sent = result.get("unique_sent")
                if total_unique is None or unique_inbox is None or unique_sent is None:
                    try:
                        lines = [
                            ln.strip()
                            for ln in full_msg.splitlines()
                            if ln.strip()
                        ]

                        def _last_int(line: str) -> int:
                            m = re.search(r"(\d+)\s*$", line)
                            return int(m.group(1)) if m else 0
                        total_unique = _last_int(lines[0]) if len(lines) > 0 else 0
                        unique_inbox = _last_int(lines[1]) if len(lines) > 1 else 0
                        unique_sent = _last_int(lines[2]) if len(lines) > 2 else 0
                    except Exception:
                        total_unique = 0
                        unique_inbox = 0
                        unique_sent = 0
                report_xlsx_path = save_excel_report(
                    username=username or "",
                    from_date=from_date or "",
                    to_date=to_date or "",
                    total=int(total_unique or 0),
                    inbox=int(unique_inbox or 0),
                    sent=int(unique_sent or 0),
                )
                common_report_path = save_common_excel_report(
                    username=username or "",
                    from_date=from_date or "",
                    to_date=to_date or "",
                    total=int(total_unique or 0),
                    inbox=int(unique_inbox or 0),
                    sent=int(unique_sent or 0),
                )
                last_files: list[str] = []
                for p in (report_txt_path, report_xlsx_path, common_report_path):
                    if p and os.path.exists(p):
                        last_files.append(p)
                multi_mode = bool(result.get("multi_mode"))
                if multi_mode:
                    if not isinstance(getattr(self, "_last_report_files", None), list):
                        self._last_report_files = []
                    for p in last_files:
                        if p not in self._last_report_files:
                            self._last_report_files.append(p)
                else:
                    self._last_report_files = last_files
                try:
                    self.btn_send_reports.config(
                        state=("normal" if self._last_report_files else "disabled")
                    )
                except Exception:
                    pass
                msg = full_msg.rstrip()
                if report_txt_path:
                    msg += f"\n\nTXT отчёт: {report_txt_path}"
                if report_xlsx_path:
                    msg += f"\nExcel отчёт: {report_xlsx_path}"
                if not result.get("silent"):
                    messagebox.showinfo(title, msg + "\n", parent=self)
            else:
                err_raw = result.get("error", "Неизвестная ошибка")
                trace = result.get("trace", "")
                full_for_report = err_raw + ("\n\n" + trace if trace else "")
                report_txt_path = save_full_report(
                    title="Ошибка",
                    message=full_for_report,
                    username=result.get("username"),
                    from_date=result.get("from_date"),
                    to_date=result.get("to_date"),
                )
                err_ui = (
                    "Не получилось посчитать письма.\n\n"
                    f"Причина: {err_raw}\n\n"
                    "Что можно сделать:\n"
                    "1) Проверь логин и пароль\n"
                    "2) Убедись, что IMAP включён в настройках почты\n"
                    "3) Попробуй уменьшить период\n"
                )
                if report_txt_path:
                    err_ui += f"\n\nОтчёт сохранён: {report_txt_path}"
                if not result.get("silent"):
                    messagebox.showerror("Ошибка", err_ui, parent=self)
        finally:
            if not result.get("multi_mode"):
                self._set_busy(False)
                self._save_config()

    def on_close(self) -> None:
        self._save_config()
        self.destroy()

if __name__ == "__main__":
    ensure_required_files()
    app = MailStatsUI()
    app.mainloop()

# pyinstaller --noconfirm --clean --onefile --windowed --name KOLORVRAT --add-data "mid_cache.json;." --add-data "mailsettings.json;." --add-data "imap_hosts.csv;." --add-data "accounts.csv;." --add-data "Отчеты;Отчеты" --hidden-import openpyxl --hidden-import openpyxl.cell --hidden-import openpyxl.styles --hidden-import openpyxl.writer.excel --collect-submodules openpyxl --collect-data openpyxl KOLORVRAT.py